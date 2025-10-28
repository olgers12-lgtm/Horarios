import pandas as pd
import numpy as np
import csv
from io import StringIO, BytesIO, TextIOWrapper
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

KNOWN_AREAS = {"calculo", "bodega", "surf", "hc", "ar", "montaje"}

def _read_uploaded_csv_filelike(filelike):
    if hasattr(filelike, "read"):
        content = filelike.read()
        if isinstance(content, bytes):
            text = content.decode("utf-8-sig", errors="replace")
        else:
            text = str(content)
    else:
        raise ValueError("filelike must have read()")
    reader = csv.reader(StringIO(text), delimiter=';')
    return [[cell.strip() for cell in row] for row in reader]

def try_parse_number(s):
    if s is None:
        return 0.0
    s = str(s).strip()
    if s == "":
        return 0.0
    filtered = ''.join(ch for ch in s if (ch.isdigit() or ch in ".-," ))
    filtered = filtered.replace(",", ".")
    try:
        return float(filtered)
    except:
        return 0.0

def parse_horario(filelike):
    content_bytes = None
    if hasattr(filelike, "read"):
        try:
            rows = _read_uploaded_csv_filelike(filelike)
            is_csv_style = any(len(rows)>0 and (rows[0][0].strip().lower() in KNOWN_AREAS) for _ in [0])
            if is_csv_style:
                return parse_horario_from_rows(rows)
            else:
                return parse_horario_from_rows(rows)
        except Exception:
            try:
                import pandas as pd
                try:
                    filelike.seek(0)
                except:
                    pass
                df = pd.read_excel(filelike, header=None, dtype=str)
                rows = df.fillna("").astype(str).values.tolist()
                return parse_horario_from_rows(rows)
            except Exception as e:
                raise e
    else:
        path = str(filelike)
        if path.lower().endswith(".xlsx") or path.lower().endswith(".xls"):
            df = pd.read_excel(path, header=None, dtype=str)
            rows = df.fillna("").astype(str).values.tolist()
            return parse_horario_from_rows(rows)
        else:
            with open(path, encoding='utf-8-sig') as f:
                reader = csv.reader(f, delimiter=';')
                rows = [[c.strip() for c in row] for row in reader]
            return parse_horario_from_rows(rows)

def parse_horario_from_rows(lines):
    capacities = []
    capacity_jobs = []
    hours_map = {}
    i = 0
    while i < len(lines):
        row = lines[i] if i < len(lines) else []
        if not row:
            i += 1
            continue
        first = (row[0] or "").strip()
        if first.lower() in KNOWN_AREAS:
            area = first
            raw_hours = row[2:]
            parsed_hours = []
            for h in raw_hours:
                if h == "":
                    continue
                try:
                    parsed_hours.append(int(h))
                except:
                    digits = ''.join(ch for ch in h if ch.isdigit())
                    if digits:
                        parsed_hours.append(int(digits))
            hours_map[area] = parsed_hours
            j = i + 1
            while j < len(lines):
                crow = lines[j]
                cfirst = (crow[0] or "").strip().lower()
                if cfirst in KNOWN_AREAS:
                    break
                if cfirst.startswith("capacidad") and "job" not in cfirst:
                    vals = crow[2:2+len(parsed_hours)]
                    for idx, h in enumerate(parsed_hours):
                        v = try_parse_number(vals[idx]) if idx < len(vals) else 0.0
                        capacities.append({"Area": area, "Day": "", "Hour": h, "Capacity": v})
                if ("capacidad job" in cfirst) or (cfirst.startswith("capacidad") and "job" in cfirst):
                    vals = crow[2:2+len(parsed_hours)]
                    for idx, h in enumerate(parsed_hours):
                        v = try_parse_number(vals[idx]) if idx < len(vals) else 0.0
                        capacity_jobs.append({"Area": area, "Day": "", "Hour": h, "CapacityJob": v})
                j += 1
            i = j
            continue
        i += 1

    caps_df = pd.DataFrame(capacities)
    capjobs_df = pd.DataFrame(capacity_jobs)
    if caps_df.empty:
        caps_df = pd.DataFrame(columns=["Area","Day","Hour","Capacity"])
    if capjobs_df.empty:
        capjobs_df = pd.DataFrame(columns=["Area","Day","Hour","CapacityJob"])
    return caps_df, capjobs_df

def aggregate_to_matrix(capacities_df, capacity_jobs_df):
    if capacities_df.empty and capacity_jobs_df.empty:
        return pd.DataFrame(columns=["Area","Day","Hour","Capacity","CapacityJob","Diferencia","Estado"])
    if capacities_df.empty:
        capacities_df = pd.DataFrame([{"Area":r["Area"],"Day":r.get("Day",""),"Hour":r["Hour"],"Capacity":0.0} for _,r in capacity_jobs_df.iterrows()])
    if capacity_jobs_df.empty:
        capacity_jobs_df = pd.DataFrame([{"Area":r["Area"],"Day":r.get("Day",""),"Hour":r["Hour"],"CapacityJob":0.0} for _,r in capacities_df.iterrows()])
    merged = pd.merge(capacities_df, capacity_jobs_df, on=["Area","Day","Hour"], how="outer")
    merged["Capacity"] = merged["Capacity"].fillna(0.0).astype(float)
    merged["CapacityJob"] = merged["CapacityJob"].fillna(0.0).astype(float)
    merged["Diferencia"] = merged["Capacity"] - merged["CapacityJob"]
    merged["Estado"] = np.where(merged["Diferencia"] >= 0, "OK", "SOBRECARGA")
    merged = merged.sort_values(by=["Area","Day","Hour"]).reset_index(drop=True)
    return merged

def read_jobs_csv_from_filelike(filelike):
    if hasattr(filelike, "read"):
        content = filelike.read()
        if isinstance(content, bytes):
            text = content.decode("utf-8-sig", errors="replace")
        else:
            text = str(content)
        df = pd.read_csv(StringIO(text))
    else:
        df = pd.read_csv(str(filelike))
    cols = {c.strip().lower(): c for c in df.columns}
    required = ["area","starthour","duration","quantity"]
    for r in required:
        if r not in cols:
            raise ValueError(f"Jobs CSV debe contener columna '{r}' (case-insensitive). Columnas encontradas: {list(df.columns)}")
    jobs = []
    for _, row in df.iterrows():
        jobid = row.get(cols.get("jobid", ""), "")
        area = str(row[cols["area"]]).strip()
        day = str(row[cols["day"]]).strip() if "day" in cols else ""
        start = int(float(row[cols["starthour"]])) if row[cols["starthour"]] not in (None,"") else 0
        duration = int(float(row[cols["duration"]])) if row[cols["duration"]] not in (None,"") else 1
        qty = float(row[cols["quantity"]]) if row[cols["quantity"]] not in (None,"") else 0.0
        jobs.append({"JobID": jobid, "Area": area, "Day": day, "StartHour": start, "Duration": duration, "Quantity": qty})
    return jobs

def assign_jobs_greedy(merged_df, jobs):
    cap = defaultdict(lambda: defaultdict(float))
    for _, r in merged_df.iterrows():
        cap[r["Area"]][int(r["Hour"])] = float(r["Capacity"])
    assignments = []
    jobs_results = []
    for job in jobs:
        jid = job["JobID"]
        area = job["Area"]
        start = job["StartHour"]
        dur = job["Duration"]
        qty = job["Quantity"]
        hours = [start + k for k in range(dur)]
        total_avail = sum(cap[area].get(h, 0.0) for h in hours)
        to_assign = min(qty, total_avail)
        assigned = 0.0
        for h in hours:
            if to_assign <= 0:
                break
            avail = cap[area].get(h, 0.0)
            alloc = min(avail, to_assign)
            if alloc > 0:
                assignments.append({"JobID": jid, "Area": area, "Day": job.get("Day",""), "Hour": h, "AssignedQuantity": alloc})
                cap[area][h] = cap[area].get(h, 0.0) - alloc
                to_assign -= alloc
                assigned += alloc
        if assigned == qty:
            status = "Assigned"
        elif assigned > 0:
            status = "Partial"
        else:
            status = "Unassigned"
        jobs_results.append({"JobID": jid, "Area": area, "Day": job.get("Day",""), "StartHour": start, "Duration": dur, "RequestedQuantity": qty, "AssignedTotal": assigned, "Status": status})
    caprem_rows = []
    for area, hours_map in cap.items():
        for h, rem in hours_map.items():
            caprem_rows.append({"Area": area, "Hour": h, "CapacityRemaining": rem})
    import pandas as pd
    return pd.DataFrame(assignments), pd.DataFrame(jobs_results), pd.DataFrame(caprem_rows)

def create_report_xlsx(merged_df, assignments_df, jobs_result_df, caprem_df):
    from io import BytesIO
    import pandas as pd
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as ew:
        merged_df.to_excel(ew, sheet_name="Summary", index=False)
        merged_df[merged["Diferencia"] < 0].to_excel(ew, sheet_name="Sobrecargas", index=False)
        merged_df.to_excel(ew, sheet_name="Unpivot", index=False)
        if not assignments_df.empty:
            assignments_df.to_excel(ew, sheet_name="Assignment", index=False)
        if not jobs_result_df.empty:
            jobs_result_df.to_excel(ew, sheet_name="JobsResult", index=False)
        if not caprem_df.empty:
            caprem_df.to_excel(ew, sheet_name="CapacityRemaining", index=False)
    buf.seek(0)
    try:
        wb = openpyxl.load_workbook(buf)
        ws = wb["Summary"]
        header = [cell.value for cell in ws[1]]
        if "Diferencia" in header:
            col_idx = header.index("Diferencia") + 1
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            rule = CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill)
            last_row = ws.max_row
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{last_row}"
            ws.conditional_formatting.add(cell_range, rule)
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out.read()
    except Exception:
        buf.seek(0)
        return buf.read()